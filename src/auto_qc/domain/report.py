"""质检报告 Excel 生成"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SEVERITY_FILLS = {
    "高": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "中": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "低": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def write_report(
    output_path: str,
    qc_results: list[dict],
    stats: dict,
) -> None:
    """生成 2-Sheet 质检报告 Excel。"""
    wb = openpyxl.Workbook()

    # Sheet 1: 合规检测
    ws_qc = wb.active
    ws_qc.title = "合规检测"
    ws_qc.sheet_properties.tabColor = "4472C4"

    headers = ["id", "时间", "意向结果", "违规规则", "问题类型", "危害程度", "证据片段", "改进建议"]
    for col, h in enumerate(headers, 1):
        _style_header(ws_qc.cell(1, col, h))

    row_num = 2
    for record in qc_results:
        violations = record.get("violations", [])
        if not violations:
            ws_qc.cell(row_num, 1, record.get("id", ""))
            ws_qc.cell(row_num, 2, record.get("time", ""))
            ws_qc.cell(row_num, 3, record.get("intent", ""))
            ws_qc.cell(row_num, 4, "通过")
            row_num += 1
        else:
            for v in violations:
                ws_qc.cell(row_num, 1, record.get("id", ""))
                ws_qc.cell(row_num, 2, record.get("time", ""))
                ws_qc.cell(row_num, 3, record.get("intent", ""))
                ws_qc.cell(row_num, 4, v.get("rule_id", ""))
                ws_qc.cell(row_num, 5, v.get("rule_name", ""))
                ws_qc.cell(row_num, 6, v.get("severity", ""))
                ws_qc.cell(row_num, 7, v.get("evidence", ""))
                ws_qc.cell(row_num, 8, v.get("suggestion", ""))
                severity = v.get("severity", "")
                if severity in SEVERITY_FILLS:
                    ws_qc.cell(row_num, 6).fill = SEVERITY_FILLS[severity]
                row_num += 1

    # Sheet 2: 统计概览
    ws_stats = wb.create_sheet("统计概览")
    ws_stats.sheet_properties.tabColor = "FFD93D"

    ws_stats.cell(1, 1, "指标")
    ws_stats.cell(1, 2, "数值")
    _style_header(ws_stats.cell(1, 1))
    _style_header(ws_stats.cell(1, 2))

    row_num = 2
    for key, label in [("total", "总对话数"), ("pass", "通过数"), ("violation_rate", "违规率")]:
        if key in stats:
            ws_stats.cell(row_num, 1, label)
            ws_stats.cell(row_num, 2, stats[key])
            row_num += 1

    row_num += 1
    rule_header = ["规则ID", "规则名称", "命中次数", "占比"]
    for col, h in enumerate(rule_header, 1):
        _style_header(ws_stats.cell(row_num, col, h))
    row_num += 1

    rules_hit = stats.get("rules_hit", {})
    rule_names = stats.get("rule_names", {})
    total_violations = sum(rules_hit.values()) if rules_hit else 0

    for rule_id in sorted(rules_hit.keys()):
        count = rules_hit[rule_id]
        pct = f"{count / total_violations * 100:.1f}%" if total_violations > 0 else "0.0%"
        ws_stats.cell(row_num, 1, rule_id)
        ws_stats.cell(row_num, 2, rule_names.get(rule_id, ""))
        ws_stats.cell(row_num, 3, count)
        ws_stats.cell(row_num, 4, pct)
        row_num += 1

    ws_stats.cell(row_num, 1, "合计")
    ws_stats.cell(row_num, 3, total_violations)
    ws_stats.cell(row_num, 4, "100.0%")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def verify_report_exists(output_path: str) -> bool:
    """验证报告文件是否生成且非空。"""
    p = Path(output_path)
    return p.exists() and p.stat().st_size > 0
